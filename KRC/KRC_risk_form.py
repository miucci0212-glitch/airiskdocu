import os
import openpyxl

def create_blank_form(input_path, output_path):
    """
    위험성평가표 양식 파일에서 작성 예시 데이터를 제거하고 
    포맷과 스타일만 유지된 깨끗한 빈 양식을 생성합니다.
    최초/정기 양식과 수시 양식의 차이를 감지하여 제목과 J열(검토/추록) 라벨을 자동으로 알맞게 적용합니다.
    """
    if not os.path.exists(input_path):
        print(f"오류: 입력 파일이 존재하지 않습니다: {input_path}")
        return

    try:
        # openpyxl로 워크북 로드 (수식 및 셀 서식 유지를 위해 data_only=False)
        wb = openpyxl.load_workbook(input_path, data_only=False)
        sheet = wb.active
        
        # 1. 상단 현장명 및 관리정보 초기화 (라벨은 유지하고 값만 비움)
        sheet['B1'] = ""  # 현장명
        sheet['B2'] = ""  # 작성일
        sheet['B4'] = ""  # 작성자
        sheet['B5'] = ""  # 관리기간
        
        # 파일 유형 분석 (파일명 기준)
        filename = os.path.basename(input_path)
        is_susi = "수시" in filename
        is_chocho_jungki = "최초" in filename or "정기" in filename
        
        # D1 셀의 문서 대제목 자동 변경
        if is_susi:
            sheet['D1'] = "수시 위험성평가서"
        elif is_chocho_jungki:
            sheet['D1'] = "최초/정기 위험성평가서"
            
        # 2. 본문 위험성평가 테이블의 예시 데이터 행 초기화 (10행부터 15행까지)
        for r in range(10, 16):
            is_even_row = (r % 2 == 1)  # 11, 13, 15행 (하단/검토 행)
            
            for c in range(1, 19):  # A열(1)부터 R열(18)까지
                cell = sheet.cell(row=r, column=c)
                
                # 읽기 전용 병합 셀(MergedCell)인 경우 건너뜀
                if type(cell).__name__ == 'MergedCell':
                    continue
                    
                # 셀 스타일은 유지하고 데이터만 삭제
                cell.value = None
            
            # 수시 위험성평가서인 경우, 하단 행의 J열(10번째 열)에 "검토/추록" 라벨을 정확하게 강제 입력합니다.
            if is_susi and is_even_row:
                sheet.cell(row=r, column=10).value = "검토/추록"
                
        # 새 파일로 저장
        wb.save(output_path)
        print(f"성공: 빈 양식 생성 완료 -> {os.path.basename(output_path)}")
        
    except PermissionError:
        print(f"\n[저장 실패] 파일이 이미 열려 있습니다: {os.path.basename(output_path)}")
        print("   -> 해당 엑셀 파일을 닫으신 후 다시 스크립트를 실행해 주세요.\n")
    except Exception as e:
        print(f"오류 발생: {str(e)}")

def main():
    # 실행 경로 기준 파일 자동 검색 및 변환
    current_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in locals() else os.getcwd()
    
    # 디렉토리 내 위험성평가표 양식 파일 탐색
    # 규칙: '위험성평가표_양식'으로 시작하고, '_빈양식'은 제외하며, 엑셀 파일인 것
    all_files = os.listdir(current_dir)
    target_files = []
    
    for f in all_files:
        if (f.startswith("위험성평가표_양식") and 
            f.endswith(".xlsx") and 
            not f.endswith("_빈양식.xlsx") and 
            not f.startswith("~$")):
            target_files.append(f)
            
    if not target_files:
        print("변환 대상 파일을 찾지 못했습니다. 폴더에 '위험성평가표_양식...' 파일이 있는지 확인해 주세요.")
        return
        
    print(f"총 {len(target_files)}개의 양식 파일을 발견했습니다. 변환을 시작합니다.\n")
    
    for filename in target_files:
        input_path = os.path.join(current_dir, filename)
        # '위험성평가표_양식_수시.xlsx' -> '위험성평가표_양식_수시_빈양식.xlsx'
        name_part, ext = os.path.splitext(filename)
        output_filename = f"{name_part}_빈양식{ext}"
        output_path = os.path.join(current_dir, output_filename)
        
        create_blank_form(input_path, output_path)

if __name__ == "__main__":
    main()
