"""위험성평가서 템플릿에 평가 결과 데이터를 채운다."""
import os
import shutil
import yaml
from copy import copy
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string

from models import AssessRequest, AssessRow


def _load_map(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _write_merged(ws, cell_range: str, value):
    ref = cell_range.split(":")[0]
    ws[ref].value = value


def fill_template(
    request: AssessRequest,
    rows: list[AssessRow],
    template_path: str,
    output_path: str,
    cell_map_path: str = "template/cell_map.yaml",
) -> str:
    cm = _load_map(cell_map_path)
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    _write_merged(ws, cm["site_name_cell"], request.site_name)
    _write_merged(ws, cm["vendor_trade_cell"], f"{request.vendor} / {request.trade}")
    period = request.period
    period_str = f"{period.start.strftime('%Y. %m. %d.')} ~ {period.end.strftime('%Y. %m. %d.')}"
    _write_merged(ws, cm["period_cell"], period_str)
    _write_merged(ws, cm["headcount_cell"], f"{request.headcount}명")
    _write_merged(ws, cm["machinery_cell"], request.machinery)
    leaders_str = f"{request.leader} / {', '.join(request.workers)}" if request.workers else request.leader
    _write_merged(ws, cm["leader_workers_cell"], leaders_str)
    equipment_str = ", ".join(request.equipment) if request.equipment else "해당없음"
    _write_merged(ws, cm["equipment_cell"], equipment_str)

    start_row = cm["body_start_row"]
    hazard_start = cm["hazard_cols"][0]
    control_start = cm["control_cols"][0]
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row_data in enumerate(rows[:12]):
        excel_row = start_row + i
        loc_col = column_index_from_string(cm["location_col"])
        work_col = column_index_from_string(cm["work_col"])

        ws.cell(row=excel_row, column=loc_col, value=row_data.location).alignment = center
        ws.cell(row=excel_row, column=work_col, value=row_data.work).alignment = center
        hazard_cell = ws[f"{hazard_start}{excel_row}"]
        hazard_cell.value = row_data.hazard
        hazard_cell.alignment = left_wrap
        control_cell = ws[f"{control_start}{excel_row}"]
        control_cell.value = row_data.control
        control_cell.alignment = left_wrap
        note_col = column_index_from_string(cm["note_col"])
        ws.cell(row=excel_row, column=note_col, value=row_data.note)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path
