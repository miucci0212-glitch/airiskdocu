"""openpyxl로 위험성평가서 양식 템플릿 생성."""
import os
import yaml
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, GradientFill
)
from openpyxl.utils import column_index_from_string


_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
MEDIUM_BORDER = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
LABEL_FILL = PatternFill("solid", fgColor="EEF1F8")


def _load_map(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _merge_and_style(ws, cell_range: str, value=None, font=None,
                     fill=None, alignment=None, border=None):
    if ":" in cell_range:
        ws.merge_cells(cell_range)
        cell_ref = cell_range.split(":")[0]
    else:
        cell_ref = cell_range
    cell = ws[cell_ref]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def build_template(output_path: str, cell_map_path: str = "template/cell_map.yaml") -> None:
    cm = _load_map(cell_map_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "위험성평가서"

    for col_letter, width in cm["column_widths"].items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 35
    for r in range(2, 5):
        ws.row_dimensions[r].height = cm["row_heights"]["header"]
    for r in range(6, 18):
        ws.row_dimensions[r].height = cm["row_heights"]["data"]
    ws.row_dimensions[18].height = cm["row_heights"]["footer"]

    title_font = Font(name="맑은 고딕", size=16, bold=True)
    label_font = Font(name="맑은 고딕", size=9, bold=True)
    header_font = Font(name="맑은 고딕", size=9, bold=True)
    normal_font = Font(name="맑은 고딕", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _merge_and_style(ws, cm["logo_range"],
                     value=f"{cm['logo_line1']}\n{cm['logo_line2']}",
                     font=Font(name="맑은 고딕", size=10, bold=True, color="1F4E79"),
                     alignment=center, border=MEDIUM_BORDER)
    _merge_and_style(ws, cm["title_range"],
                     value=cm["title_text"],
                     font=title_font, alignment=center, border=MEDIUM_BORDER)

    header_labels = [
        (cm["site_name_label_cell"], "현  장  명"),
        (cm["vendor_label_cell"], "업체명/공종"),
        (cm["period_label_cell"], "작업기간"),
        (cm["headcount_label_cell"], "작업인원"),
        (cm["machinery_label_cell"], "건설기계 종류 및 댓 수"),
        (cm["leaders_label_cell"], "작업 책임자/근로자"),
        (cm["equipment_label_cell"], "기계/기구 및 위험물질"),
    ]
    for cell_ref, label_text in header_labels:
        _merge_and_style(ws, cell_ref, value=label_text,
                         font=label_font, fill=LABEL_FILL,
                         alignment=center, border=THIN_BORDER)

    value_cells = [
        cm["site_name_cell"], cm["vendor_trade_cell"], cm["period_cell"],
        cm["headcount_cell"], cm["machinery_cell"],
        cm["leader_workers_cell"], cm["equipment_cell"],
    ]
    for cell_range in value_cells:
        _merge_and_style(ws, cell_range, font=normal_font,
                         alignment=center, border=THIN_BORDER)

    col_headers_map = cm["col_headers"]
    col_header_texts = {
        "number": "번\n호",
        "location": "작업장소",
        "work": "작업내용",
        "hazard": "위    험    요    인",
        "control": "안전보건추진계획",
        "note": "비\n고",
    }
    for key, text in col_header_texts.items():
        _merge_and_style(ws, col_headers_map[key], value=text,
                         font=header_font, fill=HEADER_FILL,
                         alignment=center, border=THIN_BORDER)

    hazard_start = cm["hazard_cols"][0]
    hazard_end = cm["hazard_cols"][-1]
    control_start = cm["control_cols"][0]
    control_end = cm["control_cols"][-1]

    for row_num in range(1, 13):
        excel_row = cm["body_start_row"] + row_num - 1
        cell = ws.cell(row=excel_row, column=1, value=row_num)
        cell.font = normal_font
        cell.alignment = center
        cell.border = THIN_BORDER
        ws.cell(row=excel_row, column=column_index_from_string(cm["location_col"])).border = THIN_BORDER
        ws.cell(row=excel_row, column=column_index_from_string(cm["location_col"])).alignment = center
        ws.cell(row=excel_row, column=column_index_from_string(cm["location_col"])).font = normal_font
        ws.cell(row=excel_row, column=column_index_from_string(cm["work_col"])).border = THIN_BORDER
        ws.cell(row=excel_row, column=column_index_from_string(cm["work_col"])).alignment = center
        ws.cell(row=excel_row, column=column_index_from_string(cm["work_col"])).font = normal_font
        ws.merge_cells(f"{hazard_start}{excel_row}:{hazard_end}{excel_row}")
        ws[f"{hazard_start}{excel_row}"].border = THIN_BORDER
        ws[f"{hazard_start}{excel_row}"].alignment = left_wrap
        ws[f"{hazard_start}{excel_row}"].font = normal_font
        ws.merge_cells(f"{control_start}{excel_row}:{control_end}{excel_row}")
        ws[f"{control_start}{excel_row}"].border = THIN_BORDER
        ws[f"{control_start}{excel_row}"].alignment = left_wrap
        ws[f"{control_start}{excel_row}"].font = normal_font
        ws.cell(row=excel_row, column=column_index_from_string(cm["note_col"])).border = THIN_BORDER
        ws.cell(row=excel_row, column=column_index_from_string(cm["note_col"])).alignment = center
        ws.cell(row=excel_row, column=column_index_from_string(cm["note_col"])).font = normal_font

    footer_texts = {
        "author": "■ 작성자(작업 책임자) :",
        "workers": "■ 근로자 :",
        "safety_manager": "■ 안전관리자 :",
        "supervisor": "■ 관리감독자 :",
        "site_manager": "■ 현장소장 :",
    }
    for key, text in footer_texts.items():
        _merge_and_style(ws, cm["footer_fields"][key], value=text,
                         font=normal_font, alignment=center, border=THIN_BORDER)

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"템플릿 생성 완료: {output_path}")


if __name__ == "__main__":
    from config import settings
    build_template(settings.template_xlsx_path, settings.cell_map_path)
