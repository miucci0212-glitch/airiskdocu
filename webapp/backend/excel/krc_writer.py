"""농어촌공사 위험성평가 양식 채우기."""
import math
import os
import shutil
from datetime import date
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import KrcMetadata, KrcRow


BODY_START_ROW = 10
ROWS_PER_ENTRY = 2
MAX_ENTRIES = 5
# 페이지 2+에서 삭제할 상단 행 수 (현장명/작성일/결재/작성자/관리기간/빈줄)
HEADER_ROWS_TO_STRIP = 7


def _format_date(d: Optional[date]) -> str:
    if d is None:
        return ""
    return d.strftime("%Y. %m. %d.")


def _template_for(krc_type: str) -> str:
    return "krc_수시_template.xlsx" if krc_type == "수시" else "krc_초기정기_template.xlsx"


def _fill_sheet(
    ws: Worksheet,
    metadata: KrcMetadata,
    rows: list[KrcRow],
    body_start_row: int = BODY_START_ROW,
    include_metadata: bool = True,
) -> None:
    """단일 시트에 메타데이터 + 최대 MAX_ENTRIES개 행을 기록한다.

    페이지 2+에서는 상단 메타데이터 영역이 잘려나가므로 include_metadata=False로
    호출하고 body_start_row를 보정해서 컬럼 헤더 직후부터 데이터를 채운다.
    """
    if include_metadata:
        ws["B1"] = metadata.site_name
        ws["B2"] = _format_date(metadata.write_date)
        ws["B4"] = metadata.approver_construction
        ws["B5"] = f"{_format_date(metadata.period_start)} ~ {_format_date(metadata.period_end)}"
        # B3, B6 are placeholder labels in the template ("(위험성평가 회의일 또는 이전)",
        # "(위험성평가 실시규정에 정해진 주기)") — leave them as-is.
        ws["M2"] = metadata.approver_construction
        ws["N2"] = metadata.approver_safety
        ws["O2"] = metadata.approver_site_manager
        ws["R2"] = metadata.inspector_supervisor

    for i, row in enumerate(rows[:MAX_ENTRIES]):
        top = body_start_row + i * ROWS_PER_ENTRY
        bot = top + 1
        ws[f"A{top}"] = row.detail_work or ""
        ws[f"A{bot}"] = row.work_location or ""
        ws[f"B{top}"] = row.equipment or ""
        ws[f"C{top}"] = row.hazard or ""
        ws[f"F{top}"] = row.accident_type or ""
        if row.frequency is not None:
            ws[f"G{top}"] = row.frequency
        if row.severity is not None:
            ws[f"H{top}"] = row.severity
        ws[f"I{top}"] = row.risk_grade or ""
        ws[f"J{top}"] = row.controls or ""
        ws[f"P{top}"] = row.improved_risk or ""
        ws[f"P{bot}"] = row.improvement_due or ""
        ws[f"R{top}"] = row.executor or ""
        ws[f"R{bot}"] = row.verifier or ""


def fill_krc_template(
    metadata: KrcMetadata,
    rows: list[KrcRow],
    template_dir: str,
    output_path: str,
) -> str:
    """행이 MAX_ENTRIES(3)를 넘으면 같은 시트를 늘리지 않고 시트를 복제해 페이지를 추가한다.
    템플릿 시트의 행 높이(75.2)가 보존되므로 4행부터 높이가 깨지지 않는다."""
    template_path = os.path.join(template_dir, _template_for(metadata.krc_type))
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    num_pages = max(1, math.ceil(len(rows) / MAX_ENTRIES))
    base_ws = wb.active
    base_title = base_ws.title

    # 시트 1은 기존 활성 시트. 추가 페이지가 필요하면 빈 상태 그대로 복제해서 데이터 채우기 직전에 sheets 리스트를 완성한다.
    sheets: list[Worksheet] = [base_ws]
    for page_idx in range(2, num_pages + 1):
        new_ws = wb.copy_worksheet(base_ws)
        suffix = f" ({page_idx})"
        max_title_len = 31 - len(suffix)
        new_ws.title = f"{base_title[:max_title_len]}{suffix}"
        sheets.append(new_ws)

    if num_pages > 1:
        suffix_first = " (1)"
        max_title_len = 31 - len(suffix_first)
        base_ws.title = f"{base_title[:max_title_len]}{suffix_first}"

    for page_idx, ws in enumerate(sheets):
        chunk = rows[page_idx * MAX_ENTRIES : (page_idx + 1) * MAX_ENTRIES]
        if page_idx == 0:
            _fill_sheet(ws, metadata, chunk)
        else:
            # 페이지 2부터는 상단 메타 영역을 시각적으로 제거한다.
            # openpyxl의 delete_rows는 병합 셀을 제대로 정리하지 못해 데이터가
            # 어긋나므로, 대신 1~HEADER_ROWS_TO_STRIP행의 높이를 0으로 만들어
            # 행을 접고 인쇄/표시상 보이지 않게 한다.
            for r in range(1, HEADER_ROWS_TO_STRIP + 1):
                ws.row_dimensions[r].height = 0
                ws.row_dimensions[r].hidden = True
            _fill_sheet(ws, metadata, chunk, include_metadata=False)

    wb.save(output_path)
    return output_path
