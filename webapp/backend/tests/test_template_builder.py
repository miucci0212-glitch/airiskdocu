import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
from openpyxl import load_workbook


def test_build_template_creates_file(tmp_path):
    from excel.template_builder import build_template
    out = tmp_path / "template.xlsx"
    build_template(str(out), cell_map_path="template/cell_map.yaml")
    assert out.exists()


def test_template_has_one_sheet(tmp_path):
    from excel.template_builder import build_template
    out = tmp_path / "template.xlsx"
    build_template(str(out), cell_map_path="template/cell_map.yaml")
    wb = load_workbook(str(out))
    assert len(wb.sheetnames) == 1


def test_template_has_title_text(tmp_path):
    from excel.template_builder import build_template
    out = tmp_path / "template.xlsx"
    build_template(str(out), cell_map_path="template/cell_map.yaml")
    wb = load_workbook(str(out))
    ws = wb.active
    title_cell = ws["C1"]
    assert title_cell.value is not None
    assert "위험성평가서" in str(title_cell.value)


def test_template_has_12_data_rows(tmp_path):
    from excel.template_builder import build_template
    out = tmp_path / "template.xlsx"
    build_template(str(out), cell_map_path="template/cell_map.yaml")
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.max_row >= 18
