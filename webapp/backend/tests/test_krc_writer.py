"""KRC 위험성평가서 writer 테스트 — 대책(J열) 셀 내 줄바꿈 정규화 검증."""
import os
from datetime import date

from openpyxl import load_workbook

from excel.krc_writer import fill_krc_template
from models import KrcMetadata, KrcRow

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "template")


def _metadata() -> KrcMetadata:
    return KrcMetadata(
        krc_type="수시",
        site_name="테스트 현장",
        write_date=date(2026, 7, 5),
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 31),
    )


def test_controls_pipe_joined_becomes_newlines(tmp_path):
    """DB 인제스트에서 ' | '로 조인된 대책이 셀 내 줄바꿈으로 변환된다."""
    output_path = str(tmp_path / "out.xlsx")
    rows = [KrcRow(controls="대책1 | 대책2 | 대책3")]

    fill_krc_template(_metadata(), rows, TEMPLATE_DIR, output_path)

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["J10"].value == "대책1\n대책2\n대책3"


def test_controls_existing_newlines_preserved(tmp_path):
    """이미 줄바꿈이 들어있는 대책은 그대로 보존된다."""
    output_path = str(tmp_path / "out.xlsx")
    rows = [KrcRow(controls="대책 A\n대책 B")]

    fill_krc_template(_metadata(), rows, TEMPLATE_DIR, output_path)

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["J10"].value == "대책 A\n대책 B"
