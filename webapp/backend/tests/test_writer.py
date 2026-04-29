import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
from openpyxl import load_workbook
from models import AssessRequest, AssessRow, Period
from datetime import date


SAMPLE_REQUEST = AssessRequest(
    site_name="서면 어반센트 데시앙 신축공사",
    vendor="㈜한창테크",
    trade="금속",
    period=Period(start=date(2026, 4, 28), end=date(2026, 4, 28)),
    headcount=2,
    leader="김재한",
    workers=["김영열"],
    equipment=["핸드그라인더", "용접기"],
    machinery="해당없음",
    locations=["현장출입구"],
    work_description="출입구 도어 시공",
)

SAMPLE_ROWS = [
    AssessRow(
        location="현장출입구",
        work="출입구 도어 시공",
        hazard="그라인더 보호덮개 미설치 상태에서 날 파단 비래",
        control="- 날접촉방지 보호덮개 설치\n- 그라인더 사용전 30초간 시운전 실시",
    ),
    AssessRow(
        location="현장출입구",
        work="출입구 도어 시공",
        hazard="용접기 홀더선 피복 손상으로 감전",
        control="- 용접홀더 피복상태 확인\n- 손상된 부분 절연테이프 테이핑",
    ),
]


@pytest.fixture
def template_path(tmp_path):
    from excel.template_builder import build_template
    t = tmp_path / "template.xlsx"
    build_template(str(t), cell_map_path="template/cell_map.yaml")
    return str(t)


def test_fill_creates_output(template_path, tmp_path):
    from excel.writer import fill_template
    out = str(tmp_path / "output.xlsx")
    fill_template(SAMPLE_REQUEST, SAMPLE_ROWS, template_path, out, "template/cell_map.yaml")
    assert os.path.exists(out)


def test_fill_writes_site_name(template_path, tmp_path):
    from excel.writer import fill_template
    import yaml
    out = str(tmp_path / "output.xlsx")
    fill_template(SAMPLE_REQUEST, SAMPLE_ROWS, template_path, out, "template/cell_map.yaml")
    wb = load_workbook(out)
    ws = wb.active
    with open("template/cell_map.yaml", encoding="utf-8") as f:
        cm = yaml.safe_load(f)
    site_cell = cm["site_name_cell"].split(":")[0]
    assert ws[site_cell].value == SAMPLE_REQUEST.site_name


def test_fill_writes_data_rows(template_path, tmp_path):
    from excel.writer import fill_template
    import yaml
    out = str(tmp_path / "output.xlsx")
    fill_template(SAMPLE_REQUEST, SAMPLE_ROWS, template_path, out, "template/cell_map.yaml")
    wb = load_workbook(out)
    ws = wb.active
    with open("template/cell_map.yaml", encoding="utf-8") as f:
        cm = yaml.safe_load(f)
    start_row = cm["body_start_row"]
    hazard_col = cm["hazard_cols"][0]
    cell_val = ws[f"{hazard_col}{start_row}"].value
    assert cell_val is not None
    assert "그라인더" in str(cell_val)
