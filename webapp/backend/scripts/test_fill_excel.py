import os
from datetime import date
from models import KrcMetadata, KrcRow
from excel.krc_writer import fill_krc_template

def test_fill():
    metadata = KrcMetadata(
        krc_type="최초/정기",
        site_name="테스트 현장명",
        write_date=date(2026, 5, 21),
        writer="홍길동",
        period_start=date(2026, 5, 21),
        period_end=date(2026, 6, 20),
        approver_construction="공사결재자",
        approver_safety="안전결재자",
        approver_site_manager="소장결재자",
        inspector_supervisor="감독점검자"
    )
    
    rows = [
        KrcRow(
            detail_work="세부작업 1",
            work_location="위치 1",
            equipment="장비 1",
            hazard="유해위험요인 1",
            accident_type="재해형태 1",
            frequency=2,
            severity=3,
            risk_grade="상",
            controls="예방대책 1",
            improved_risk="1/2 (하)",
            improvement_due="2026-05-31",
            executor="이행담당 1",
            verifier="확인담당 1"
        )
    ]
    
    template_dir = "/Users/kimmiso/Desktop/위험성평가 모델/webapp/backend/template"
    output_path = "/Users/kimmiso/Desktop/위험성평가 모델/webapp/backend/template/test_output.xlsx"
    
    try:
        fill_krc_template(metadata, rows, template_dir, output_path)
        print("Success! Test excel generated at:", output_path)
        
        # Verify the cell values
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        print("B1 (Site Name):", ws["B1"].value)
        print("B2 (Write Date):", ws["B2"].value)
        print("B4 (Writer):", ws["B4"].value)
        print("B5 (Period):", ws["B5"].value)
        print("M2 (Approver Construction):", ws["M2"].value)
        print("A10 (Detail Work):", ws["A10"].value)
        print("A11 (Work Location):", ws["A11"].value)
        print("C10 (Hazard):", ws["C10"].value)
    except Exception as e:
        print("Error during fill_krc_template:", e)

if __name__ == "__main__":
    test_fill()
